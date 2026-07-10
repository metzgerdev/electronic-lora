"""Sync edits from dataset/review_sheet.xlsx back into dataset.core.json.

Matches rows to samples by filename. Applies: caption, bpm, key(scale),
labeled. Rows with verdict=exclude are removed from the dataset (reported).
vocal_status downgraded to anything containing "vocal"/"full" also excludes.
The JSON stays the training source of truth; run this after each editing
session. A backup of the previous JSON is kept as dataset.core.json.bak.
"""

import json
import shutil
from pathlib import Path

from openpyxl import load_workbook

DATASET = Path(__file__).parent.parent / "dataset"


def main():
    json_path = DATASET / "dataset.core.json"
    data = json.loads(json_path.read_text())
    ws = load_workbook(DATASET / "review_sheet.xlsx")["review"]

    header = [c.value for c in ws[1]]
    rows = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(header, r))
        if row.get("filename"):
            rows[row["filename"]] = row

    kept, excluded, edited = [], [], 0
    for s in data["samples"]:
        row = rows.get(s["filename"])
        if row is None:
            kept.append(s)  # not in sheet -> untouched
            continue
        verdict = str(row.get("verdict") or "keep").strip().lower()
        vstat = str(row.get("vocal_status") or "").strip().lower()
        if verdict == "exclude" or "full" in vstat:
            excluded.append(s["filename"])
            continue

        changed = False
        if row.get("caption") and row["caption"] != s["caption"]:
            s["caption"] = str(row["caption"]).strip()
            changed = True
        if row.get("bpm") and int(row["bpm"]) != s.get("bpm"):
            s["bpm"] = int(row["bpm"])
            changed = True
        if row.get("key") and str(row["key"]).strip() != s.get("keyscale"):
            s["keyscale"] = str(row["key"]).strip()
            changed = True
        new_labeled = str(row.get("labeled")).strip().upper() == "TRUE"
        if new_labeled != s.get("labeled"):
            s["labeled"] = new_labeled
            changed = True
        edited += changed
        kept.append(s)

    data["samples"] = kept
    data["metadata"]["num_samples"] = len(kept)
    shutil.copy(json_path, json_path.with_suffix(".json.bak"))
    json_path.write_text(json.dumps(data, indent=2))

    n_labeled = sum(1 for s in kept if s.get("labeled"))
    print(f"{len(kept)} tracks kept | {edited} edited | {n_labeled} signed off (labeled=TRUE)")
    for f in excluded:
        print(f"  EXCLUDED: {f}")


if __name__ == "__main__":
    main()
