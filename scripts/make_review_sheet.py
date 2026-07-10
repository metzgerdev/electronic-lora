"""Build dataset/review_sheet.xlsx — the editable ear-review worksheet.

Merges dataset.core.json (source of truth) with the bpm/key provenance from
captions_review.csv and vocal_status from core_review.csv. Estimated
(unverified) BPM/key cells are highlighted yellow.

Edit in any spreadsheet app, save, then run apply_review.py to sync edits
back into dataset.core.json.

Editable columns: verdict, bpm, key, vocal_status, caption, notes, labeled.
`verdict` = keep (default) | exclude — set exclude to drop a track from v1
(e.g. you discover it's actually a full-vocal track).
"""

import csv
import json
from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DATASET = Path(__file__).parent.parent / "dataset"
AUDIO_ROOT = Path.home() / "Documents/Code/training_music"

COLUMNS = [
    ("filename", 52), ("listen", 8), ("verdict", 9), ("labeled", 8), ("bpm", 6),
    ("bpm_source", 11), ("key", 10), ("key_source", 11), ("genre", 20),
    ("vocal_status", 12), ("caption", 90), ("notes", 30),
]

YELLOW = PatternFill(start_color="FFF3B0", end_color="FFF3B0", fill_type="solid")
LINK_FONT = Font(color="0563C1", underline="single")


def main():
    data = json.loads((DATASET / "dataset.core.json").read_text())
    sources = {r["filename"]: r for r in csv.DictReader(open(DATASET / "captions_review.csv"))}
    vocal = {r["filename"]: r["vocal_status"] for r in csv.DictReader(open(DATASET / "core_review.csv"))}

    wb = Workbook()
    ws = wb.active
    ws.title = "review"

    for col, (name, width) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=col, value=name)
        c.font = Font(bold=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"

    for row, s in enumerate(data["samples"], 2):
        fn = s["filename"]
        src = sources.get(fn, {})
        values = {
            "filename": fn,
            "listen": "▶ play",
            "verdict": "keep",
            "labeled": "TRUE" if s.get("labeled") else "FALSE",
            "bpm": s.get("bpm"),
            "bpm_source": src.get("bpm_source", ""),
            "key": s.get("keyscale", ""),
            "key_source": src.get("key_source", ""),
            "genre": s.get("genre", ""),
            "vocal_status": vocal.get(fn, ""),
            "caption": s.get("caption", ""),
            "notes": "",
        }
        for col, (name, _) in enumerate(COLUMNS, 1):
            c = ws.cell(row=row, column=col, value=values[name])
            if name == "listen":
                # served by scripts/serve_audio.py (Google Sheets can't open file:// links)
                c.hyperlink = f"http://localhost:8765/{quote(fn)}"
                c.font = LINK_FONT
            if name == "caption":
                c.alignment = Alignment(wrap_text=True, vertical="top")
            if name in ("bpm", "key") and "VERIFY" in values.get(f"{name}_source", ""):
                c.fill = YELLOW

    out = DATASET / "review_sheet.xlsx"
    wb.save(out)
    print(f"wrote {out} ({len(data['samples'])} tracks; yellow = unverified estimate)")


if __name__ == "__main__":
    main()
